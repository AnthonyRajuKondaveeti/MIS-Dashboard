"""
Daily Welleazy MIS (Mon-Fri)
Comprehensive daily reporting with Drug/Non-Drug bifurcation
Following exact documentation steps 1-11
"""

import pandas as pd
import numpy as np
from datetime import datetime
from welleazy_config import MISConfig
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DailyMISGenerator:
    """Generates Daily Welleazy MIS Report"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        
        # Ensure case_id column exists
        if 'case_id' not in self.df.columns:
            logger.warning("case_id column not found, creating auto-generated IDs")
            self.df['case_id'] = range(1, len(self.df) + 1)
        
        # Standardize client_name
        if 'client_name' in self.df.columns:
            self.df['client_name'] = self.df['client_name'].astype(str).str.strip()
            # Replace 'nan' string with actual NaN or empty
            self.df['client_name'] = self.df['client_name'].replace(['nan', 'None', '', 'NaN'], np.nan)
        
        self.case_status_report = None
        self.client_merged = None
        self.drug_report = None
        self.non_drug_report = None
    
    def generate(self) -> dict:
        """
        Generate complete Daily MIS Report following Steps 1-11
        """
        logger.info("Generating Daily Welleazy MIS Report...")
        
        try:
            # --- Pre-processing ---
            # Apply Service Categorization using Config (Step 9 preparation)
            self._apply_service_categorization()
            
            # --- Step 1-4: Total Case Received (Cohort: Case Entry Date) ---
            # We pass 'None' for service_filter to get the aggregate for Client Report
            received_df = self._get_received_data(service_filter=None)
            logger.info(f"Received Data: {len(received_df)} clients")
            
            # --- Step 5: Closed Data (Cohort: Case Entry Date, Status=Closed) ---
            closed_cohort_df = self._get_closed_cohort_data(service_filter=None)
            logger.info(f"Closed (Cohort) Data: {len(closed_cohort_df)} clients")
            
            # --- Step 6: Closure Data (Activity: Report Upload Date) ---
            closure_activity_df = self._get_closure_activity_data(service_filter=None)
            logger.info(f"Closure (Activity) Data: {len(closure_activity_df)} clients")
            
            # --- Step 7-8: Merge Data ---
            self.client_merged = self._merge_reports(received_df, closed_cohort_df, closure_activity_df)
            logger.info(f"Client Merged Report: {self.client_merged.shape}")
            
            # --- Step 2-3: Case Status Report ---
            # This is separate from the client table
            self.case_status_report = self._generate_case_status_report()
            logger.info(f"Case Status Report generated: {self.case_status_report.shape}")
            
            # --- Step 9: Bifurcate Drug and Non-Drug ---
            # We need to re-run the extraction/merge for specific services
            logger.info("Generating Drug Report...")
            self.drug_report = self._generate_service_specific_report('Drug')
            
            logger.info("Generating Non-Drug Report...")
            self.non_drug_report = self._generate_service_specific_report('Non-Drug')
            
            return {
                'case_status': self.case_status_report,
                'client_merged': self.client_merged,
                'drug': self.drug_report,
                'non_drug': self.non_drug_report
            }
            
        except Exception as e:
            logger.error(f"Error generating Daily MIS Report: {str(e)}", exc_info=True)
            return {
                'error': str(e)
            }

    def _apply_service_categorization(self):
        """Use MISConfig to categorize services"""
        if 'service' in self.df.columns:
            self.df['service_type'] = self.df['service'].apply(MISConfig.categorize_service)
        else:
            self.df['service_type'] = 'Non-Drug' # Default
            
    def _prepare_pivot(self, df: pd.DataFrame, date_col: str, service_filter: str = None) -> pd.DataFrame:
        """
        Generic helper to filter DF and pivot by Client x Month
        """
        if df.empty:
            return pd.DataFrame()
            
        # Ensure date column
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        
        # Filter by Date (Last 3 Month Period)
        months_back = MISConfig.LOOKBACK_MONTHS.get('daily_mis', 3)
        # Assuming we want full months including current
        start_date = (pd.Timestamp.now() - pd.DateOffset(months=months_back-1)).replace(day=1)
        
        # Filter valid dates only
        mask = df[date_col] >= start_date
        filtered = df[mask].copy()
        
        # Filter by Service if requested
        if service_filter:
            filtered = filtered[filtered['service_type'] == service_filter]
            
        if filtered.empty:
            return pd.DataFrame()

        # Create Month Column
        filtered['month_col'] = filtered[date_col].dt.to_period('M').astype(str)
        
        # Pivot: Client x Month
        pivot = pd.pivot_table(
            filtered,
            values='case_id',
            index='client_name',
            columns='month_col',
            aggfunc='count',
            fill_value=0
        )
        
        # Clean Index
        pivot.index = pivot.index.astype(str).str.strip()
        pivot = pivot[pivot.index != 'nan']
        pivot = pivot[pivot.index != 'None']
        
        return pivot

    def _get_received_data(self, service_filter=None) -> pd.DataFrame:
        """Step 1-4: Filter by Case Entry Date"""
        return self._prepare_pivot(self.df, 'case_entry_date', service_filter)

    def _get_closed_cohort_data(self, service_filter=None) -> pd.DataFrame:
        """Step 5: Filter by Case Entry Date AND Status=Closed"""
        # Filter Status
        closed_statuses = MISConfig.get_status_filter('closed_cases')
        norm_closed = [MISConfig.normalize_status_value(s) for s in closed_statuses]
        
        # Normalize status
        status_s = self.df['case_status'].apply(MISConfig.normalize_status_value)
        
        # Apply strict status filter first
        closed_df = self.df[status_s.isin(norm_closed)].copy()
        
        # Then Pivot on ENTRY Date (Cohort)
        return self._prepare_pivot(closed_df, 'case_entry_date', service_filter)

    def _get_closure_activity_data(self, service_filter=None) -> pd.DataFrame:
        """Step 6: Filter by Report Upload Date (Activity)"""
        # User step 6 says: "Go to raw dump... select latest 3 months in Report uploaded date... This data set will close data set"
        # It implies this is purely activity based on report upload, regardless of entry date.
        
        if 'report_upload_date' not in self.df.columns:
            logger.warning("report_upload_date column missing")
            return pd.DataFrame()

        df_activity = self.df[self.df['report_upload_date'].notna()].copy()
        return self._prepare_pivot(df_activity, 'report_upload_date', service_filter)

    def _merge_reports(self, received: pd.DataFrame, closed_cohort: pd.DataFrame, closure_activity: pd.DataFrame) -> pd.DataFrame:
        """
        Step 7-8: Merge tables and calculate keys
        """
        # create a union of all clients
        # Use set comprehension to avoid issues if any DF is empty (index would be empty)
        clients_set = set()
        if not received.empty: clients_set.update(received.index)
        if not closed_cohort.empty: clients_set.update(closed_cohort.index)
        if not closure_activity.empty: clients_set.update(closure_activity.index)
            
        all_clients = sorted(list(clients_set))
        if not all_clients:
            return pd.DataFrame()
        
        # Identify all relevant months from the dataframes (sorted)
        month_set = set()
        if not received.empty: month_set.update(received.columns)
        if not closed_cohort.empty: month_set.update(closed_cohort.columns)
        if not closure_activity.empty: month_set.update(closure_activity.columns)
            
        all_months = sorted(list(month_set))
        
        # Create result DF
        result = pd.DataFrame(index=all_clients)
        result.index.name = 'client_name'
        
        # Clean 'Welleazy Demo'
        if 'Welleazy Demo' in result.index:
            result = result.drop('Welleazy Demo')
        
        for month in all_months:
            # Display format (YYYY-MM to MMM-YY)
            try:
                display_month = pd.Period(month).strftime('%b-%y')
            except:
                display_month = month
                
            # 1. Received
            rec_series = received[month] if (not received.empty and month in received.columns) else pd.Series(0, index=result.index)
            rec_series = rec_series.reindex(result.index, fill_value=0)
            
            # 2. Closed (Cohort) - "Closer as per received date"
            cohort_series = closed_cohort[month] if (not closed_cohort.empty and month in closed_cohort.columns) else pd.Series(0, index=result.index)
            cohort_series = cohort_series.reindex(result.index, fill_value=0)
            
            # 3. Closure (Activity) - "Closer as per closer date"
            # Note: The month column in closure_activity_df represents the UPLOAD month, 
            # while in received definitions it represents ENTRY month.
            # We align them by the string value (e.g. Jan data in Jan column).
            act_series = closure_activity[month] if (not closure_activity.empty and month in closure_activity.columns) else pd.Series(0, index=result.index)
            act_series = act_series.reindex(result.index, fill_value=0)
            
            # 4. Percentage (Closed Cohort / Received)
            # Formula: (Closed / Case received)
            with np.errstate(divide='ignore', invalid='ignore'):
                pct_vals = (cohort_series / rec_series * 100).fillna(0)
                pct_vals = pct_vals.replace([np.inf, -np.inf], 0)
                
            # Format % as string
            pct_series = pct_vals.round(0).astype(int).astype(str) + '%'
            # Clear if no received cases (Optional, but looks cleaner)
            pct_series[rec_series == 0] = '' 
            
            # Add columns
            result[f'{display_month} - Case Received'] = rec_series
            result[f'{display_month} - Closer as per received date'] = cohort_series
            result[f'{display_month} - Closer as per closer date'] = act_series
            result[f'{display_month} - % as per received date'] = pct_series
            
        return result

    def _generate_case_status_report(self) -> pd.DataFrame:
        """Step 2-3: Case Status Report"""
        # Pivot: Status x (Month + Service)
        # Needs last 3 months filtered by Entry Date
        
        df = self.df.copy()
        
        # Filter Date
        months_back = MISConfig.LOOKBACK_MONTHS.get('daily_mis', 3)
        start_date = (pd.Timestamp.now() - pd.DateOffset(months=months_back-1)).replace(day=1)
        
        if 'case_entry_date' not in df.columns:
            return pd.DataFrame()
            
        df['case_entry_date'] = pd.to_datetime(df['case_entry_date'], errors='coerce')
        mask = df['case_entry_date'] >= start_date
        filtered = df[mask].copy()
        
        if filtered.empty:
            return pd.DataFrame()
            
        # Clean Status
        filtered['case_status'] = filtered['case_status'].fillna('Unknown').astype(str).str.strip()
        filtered = filtered[filtered['case_status'] != '']
        filtered = filtered[filtered['case_status'] != 'nan']
        
        # Month Col
        filtered['entry_month'] = filtered['case_entry_date'].dt.to_period('M').astype(str)
        
        # Pivot
        pivot = pd.pivot_table(
            filtered,
            values='case_id',
            index='case_status',
            columns=['entry_month', 'service_type'],
            aggfunc='count',
            fill_value=0
        )
        
        # Calculate Row Totals
        pivot['Total'] = pivot.sum(axis=1)
        
        # Add Column Totals row
        pivot.loc['Total'] = pivot.sum()
        
        return pivot

    def _generate_service_specific_report(self, service_type: str) -> pd.DataFrame:
        """Wrapper to generate client report filtered by service"""
        rec = self._get_received_data(service_filter=service_type)
        clo_cohort = self._get_closed_cohort_data(service_filter=service_type)
        clo_act = self._get_closure_activity_data(service_filter=service_type)
        
        return self._merge_reports(rec, clo_cohort, clo_act)

    def export_to_excel(self, filepath: str):
        """Export to Excel"""
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                if self.case_status_report is not None and not self.case_status_report.empty:
                    self.case_status_report.to_excel(writer, sheet_name='Case Status')
                    self._format_worksheet(writer.sheets['Case Status'])
                    
                if self.client_merged is not None and not self.client_merged.empty:
                    self.client_merged.to_excel(writer, sheet_name='Client Report')
                    self._format_worksheet(writer.sheets['Client Report'])
                    
                if self.drug_report is not None and not self.drug_report.empty:
                    self.drug_report.to_excel(writer, sheet_name='Drug Cases')
                    self._format_worksheet(writer.sheets['Drug Cases'])
                    
                if self.non_drug_report is not None and not self.non_drug_report.empty:
                    self.non_drug_report.to_excel(writer, sheet_name='Non-Drug Cases')
                    try:
                        self._format_worksheet(writer.sheets['Non-Drug Cases'])
                    except: pass
            logger.info(f"Exported to {filepath}")
        except Exception as e:
            logger.error(f"Export failed: {e}")

    def _format_worksheet(self, worksheet):
        """Apply formatting to worksheet"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        max_rows = worksheet.max_row
        max_cols = worksheet.max_column
        
        # Iterate over all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Header Row (Row 1 usually, sometimes 2 if multiindex)
        for row_idx in range(1, 4): # First 3 rows to be safe
            for col_idx in range(1, max_cols + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                # Apply header style
                if cell.value: # Heuristic: if cell has value in top rows, likely header
                     cell.fill = header_fill
                     cell.font = header_font
                     cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Column Widths
        for col_idx in range(1, max_cols + 1):
             worksheet.column_dimensions[get_column_letter(col_idx)].width = 20

def generate_daily_mis(df: pd.DataFrame) -> dict:
    generator = DailyMISGenerator(df)
    return generator.generate()
