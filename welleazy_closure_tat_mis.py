"""
Closure TAT and Appointment Scheduled TAT MIS Logic
Analyzes turnaround times for closed cases
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from welleazy_config import MISConfig
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ClosureTATMISGenerator:
    """Generates Closure TAT and Scheduled TAT MIS Report"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        
        # Ensure case_id column exists
        if 'case_id' not in self.df.columns:
            logger.warning("case_id column not found, creating auto-generated IDs")
            self.df['case_id'] = range(1, len(self.df) + 1)
        
        self.closure_report = None
        self.scheduled_report = None
    
    def generate(self) -> tuple:
        """
        Generate both Closure TAT and Scheduled TAT reports
        
        Returns:
            Tuple of (closure_tat_pivot, scheduled_tat_pivot)
        """
        logger.info("Generating Closure TAT and Scheduled TAT MIS Reports...")
        
        # Step 1: Filter closed cases from last N months
        filtered_df = self._filter_recent_closed_cases()
        
        if filtered_df.empty:
            logger.warning("No closed cases found in the specified period")
            return pd.DataFrame(), pd.DataFrame()
        
        # Step 2: Calculate TAT metrics
        df_with_tats = self._calculate_tats(filtered_df)
        
        # Step 3: Cap and clean TAT values
        df_clean = self._clean_tat_values(df_with_tats)
        
        # Step 4: Generate Closure TAT pivot
        closure_pivot = self._generate_closure_pivot(df_clean)
        
        # Step 5: Generate Scheduled TAT pivot
        scheduled_pivot = self._generate_scheduled_pivot(df_clean)
        
        self.closure_report = closure_pivot
        self.scheduled_report = scheduled_pivot
        
        logger.info(f"Reports generated for {len(df_clean)} closed cases")
        
        return closure_pivot, scheduled_pivot
    
    def _filter_recent_closed_cases(self) -> pd.DataFrame:
        """Filter for closed cases from last N months"""
        # Get lookback period
        months_back = MISConfig.LOOKBACK_MONTHS['closure_tat']
        cutoff_date = pd.Timestamp.now() - pd.DateOffset(months=months_back)
        
        # Ensure date column is datetime
        self.df['case_completion_date'] = pd.to_datetime(
            self.df['case_completion_date'], errors='coerce'
        )
        
        # DEBUG: Show unique status values in data
        unique_statuses = self.df['case_status'].dropna().unique()
        logger.info(f"DEBUG: Found {len(unique_statuses)} unique status values in data:")
        for status in sorted(unique_statuses):
            normalized = MISConfig.normalize_status_value(status)
            logger.info(f"  Original: '{status}' -> Normalized: '{normalized}'")
        
        # Filter by date and status - use config for consistency
        closed_statuses = MISConfig.get_status_filter('closed_cases')
        normalized_closed = [MISConfig.normalize_status_value(s) for s in closed_statuses]
        
        logger.info(f"DEBUG: Looking for normalized closed statuses: {normalized_closed}")
        
        # Normalize status values in data for matching
        status_normalized = self.df['case_status'].apply(MISConfig.normalize_status_value)
        
        # Count matches
        status_matches = status_normalized.isin(normalized_closed).sum()
        logger.info(f"DEBUG: Found {status_matches} records matching closed status")
        
        # Count date matches
        date_matches = (self.df['case_completion_date'] >= cutoff_date).sum()
        logger.info(f"DEBUG: Found {date_matches} records with case_completion_date >= {cutoff_date}")
        
        mask = (
            (self.df['case_completion_date'] >= cutoff_date) &
            (status_normalized.isin(normalized_closed))
        )
        
        filtered = self.df[mask].copy()
        logger.info(f"Filtered {len(filtered)} closed cases from last {months_back} months")
        
        return filtered
    
    def _calculate_tats(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculate Closure TAT and Scheduled TAT"""
        # Ensure all date columns are datetime
        date_cols = ['appointment_date', 'case_completion_date', 
                     'appointment_fixed_date', 'case_received_date']
        
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        # Calculate Closure TAT using NETWORKDAYS formula: NETWORKDAYS(start, end, 11) - 1
        # The 11 represents holidays (simplified - you can pass actual holiday list)
        df['closure_tat'] = df.apply(
            lambda row: self._calculate_networkdays(
                row['appointment_date'], 
                row['case_completion_date']
            ) - 1,  # Subtract 1 as per Excel formula
            axis=1
        )
        
        # Calculate Scheduled TAT (calendar days between received and appointment fixed)
        df['scheduled_tat'] = (
            df['appointment_fixed_date'] - df['case_received_date']
        ).dt.days
        
        logger.info("Calculated TAT metrics using NETWORKDAYS formula")
        
        return df
    
    def _calculate_networkdays(self, start_date: pd.Timestamp, end_date: pd.Timestamp) -> int:
        """
        Calculate working days between two dates (Monday-Friday only)
        Excludes Saturdays, Sundays, and Indian public holidays
        Implements Excel's NETWORKDAYS function
        
        Args:
            start_date: Start date
            end_date: End date
        
        Returns:
            Number of working days (Mon-Fri excluding holidays)
        """
        if pd.isna(start_date) or pd.isna(end_date):
            return 0
        
        # Define Indian public holidays for 2024-2026
        # Update this list annually with actual Indian public holidays
        indian_holidays = [
            # 2024 holidays
            '2024-01-26',  # Republic Day
            '2024-03-25',  # Holi
            '2024-03-29',  # Good Friday
            '2024-04-17',  # Ram Navami
            '2024-04-21',  # Mahavir Jayanti
            '2024-05-23',  # Buddha Purnima
            '2024-06-17',  # Eid ul-Adha
            '2024-07-17',  # Muharram
            '2024-08-15',  # Independence Day
            '2024-08-26',  # Janmashtami
            '2024-09-16',  # Eid-e-Milad
            '2024-10-02',  # Gandhi Jayanti
            '2024-10-12',  # Dussehra
            '2024-10-31',  # Diwali
            '2024-11-01',  # Diwali (Day 2)
            '2024-11-15',  # Guru Nanak Jayanti
            '2024-12-25',  # Christmas
            
            # 2025 holidays
            '2025-01-26',  # Republic Day
            '2025-03-14',  # Holi
            '2025-03-30',  # Ram Navami
            '2025-04-10',  # Mahavir Jayanti
            '2025-04-18',  # Good Friday
            '2025-05-12',  # Buddha Purnima
            '2025-06-07',  # Eid ul-Adha
            '2025-07-06',  # Muharram
            '2025-08-15',  # Independence Day
            '2025-08-16',  # Janmashtami
            '2025-09-05',  # Eid-e-Milad
            '2025-10-02',  # Gandhi Jayanti
            '2025-10-02',  # Dussehra
            '2025-10-20',  # Diwali
            '2025-11-05',  # Guru Nanak Jayanti
            '2025-12-25',  # Christmas
            
            # 2026 holidays
            '2026-01-26',  # Republic Day
            '2026-03-04',  # Holi
            '2026-03-20',  # Ram Navami
            '2026-03-30',  # Mahavir Jayanti
            '2026-04-03',  # Good Friday
            '2026-05-01',  # Buddha Purnima
            '2026-05-28',  # Eid ul-Adha
            '2026-06-26',  # Muharram
            '2026-08-05',  # Janmashtami
            '2026-08-15',  # Independence Day
            '2026-08-26',  # Eid-e-Milad
            '2026-10-02',  # Gandhi Jayanti
            '2026-10-22',  # Dussehra
            '2026-11-08',  # Diwali
            '2026-11-25',  # Guru Nanak Jayanti
            '2026-12-25',  # Christmas
        ]
        
        # Convert to numpy datetime64 format
        holidays = np.array(indian_holidays, dtype='datetime64[D]')
        
        # Custom week mask: Mon-Sat are working days (1), Sun is off (0)
        # weekmask format: Mon Tue Wed Thu Fri Sat Sun
        # User specified NETWORKDAYS(..., 11) which corresponds to Sunday-only weekend
        weekmask = [1, 1, 1, 1, 1, 1, 0]
        
        # Use numpy's busday_count with custom weekmask and holidays
        # This excludes Sundays and the specified holidays
        try:
            working_days = np.busday_count(
                start_date.date(),
                end_date.date(),
                weekmask=weekmask,
                holidays=holidays
            )
            return int(working_days)
        except:
            # Fallback if dates are invalid
            return 0
    
    def _clean_tat_values(self, df: pd.DataFrame) -> pd.DataFrame:
        """Cap and clean TAT values as per documentation"""
        # Convert negative TATs to 0
        df['closure_tat'] = df['closure_tat'].clip(lower=0)
        df['scheduled_tat'] = df['scheduled_tat'].clip(lower=0)
        
        # Cap TATs at maximum value (>=3 becomes 3)
        max_tat = MISConfig.TAT_MAX_CAP
        df['closure_tat'] = df['closure_tat'].apply(lambda x: max_tat if x >= max_tat else x)
        df['scheduled_tat'] = df['scheduled_tat'].apply(lambda x: max_tat if x >= max_tat else x)
        
        # Fill NaN with 0
        df['closure_tat'].fillna(0, inplace=True)
        df['scheduled_tat'].fillna(0, inplace=True)
        
        # Convert to int
        df['closure_tat'] = df['closure_tat'].astype(int)
        df['scheduled_tat'] = df['scheduled_tat'].astype(int)
        
        logger.info("Cleaned and capped TAT values (negative→0, >=3→3)")
        
        return df
    
    def _generate_closure_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate Closure TAT pivot table with percentages
        Structure: Client Name (rows) × (Completion Month, Closure TAT) (columns)
        """
        # Remove any None values from client_name before pivot
        df = df[df['client_name'].notna()].copy()
        
        # Create month-year column from case completion date
        df['completion_month'] = df['case_completion_date'].dt.to_period('M').astype(str)
        
        # Create pivot table: Client × (Month × TAT)
        pivot = pd.pivot_table(
            df,
            values='case_id',
            index='client_name',
            columns=['completion_month', 'closure_tat'],
            aggfunc='count',
            fill_value=0
        )
        
        # Remove any None/NaN from index
        pivot = pivot.dropna(how='all')
        pivot = pivot[pivot.index.notna()]
        
        # Ensure index has a name
        if pivot.index.name is None:
            pivot.index.name = 'client_name'
        
        # Calculate percentages for each month
        pivot_pct = self._calculate_tat_percentages(pivot)
        
        # Flatten multi-index columns to remove bucket headers
        if isinstance(pivot_pct.columns, pd.MultiIndex):
            pivot_pct.columns = [f'{month} (TAT {tat})' for month, tat in pivot_pct.columns]
        
        return pivot_pct
    
    def _generate_scheduled_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate Scheduled TAT pivot table with percentages
        Structure: Client Name (rows) × (Completion Month, Scheduled TAT) (columns)
        """
        # Remove any None values from client_name before pivot
        df = df[df['client_name'].notna()].copy()
        
        # Create month-year column
        df['completion_month'] = df['case_completion_date'].dt.to_period('M').astype(str)
        
        # Create pivot table: Client × (Month × TAT)
        pivot = pd.pivot_table(
            df,
            values='case_id',
            index='client_name',
            columns=['completion_month', 'scheduled_tat'],
            aggfunc='count',
            fill_value=0
        )
        
        # Remove any None/NaN from index
        pivot = pivot.dropna(how='all')
        pivot = pivot[pivot.index.notna()]
        
        # Ensure index has a name
        if pivot.index.name is None:
            pivot.index.name = 'client_name'
        
        # Calculate percentages
        pivot_pct = self._calculate_tat_percentages(pivot)
        
        # Flatten multi-index columns to remove bucket headers
        if isinstance(pivot_pct.columns, pd.MultiIndex):
            pivot_pct.columns = [f'{month} (TAT {tat})' for month, tat in pivot_pct.columns]
        
        return pivot_pct
    
    def _calculate_tat_percentages(self, pivot: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate percentage distribution for each TAT bucket by month
        Formula: (Individual TAT Count / Total for Month) × 100
        """
        if pivot.empty or not isinstance(pivot.columns, pd.MultiIndex):
            return pivot
        
        # Create result dataframe with same structure
        result = pivot.copy()
        
        # Get unique months from column multi-index
        months = pivot.columns.get_level_values(0).unique()
        
        for month in months:
            # Get month data
            month_data = pivot[month]
            
            # Calculate total for each client in this month
            month_totals = month_data.sum(axis=1)
            
            # Calculate percentage for each TAT bucket
            for tat in month_data.columns:
                # Avoid division by zero
                result[(month, tat)] = month_data[tat].div(month_totals).fillna(0) * 100
                result[(month, tat)] = result[(month, tat)].round(2)
        
        return result
    
    def export_to_excel(self, filepath: str):
        """Export both reports to Excel with formatting"""
        if self.closure_report is None or self.scheduled_report is None:
            logger.warning("No reports to export")
            return
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write Closure TAT report
            if not self.closure_report.empty:
                # Remove any None rows
                closure_clean = self.closure_report.copy()
                closure_clean = closure_clean[closure_clean.index.notna()]
                closure_clean = closure_clean.dropna(how='all')
                
                # Write with index (client_name will be the first column)
                closure_clean.to_excel(writer, sheet_name='Closure TAT')
                self._format_worksheet(writer.sheets['Closure TAT'], 'Closure TAT (%)')
            
            # Write Scheduled TAT report
            if not self.scheduled_report.empty:
                # Remove any None rows
                scheduled_clean = self.scheduled_report.copy()
                scheduled_clean = scheduled_clean[scheduled_clean.index.notna()]
                scheduled_clean = scheduled_clean.dropna(how='all')
                
                # Write with index (client_name will be the first column)
                scheduled_clean.to_excel(writer, sheet_name='Scheduled TAT')
                self._format_worksheet(writer.sheets['Scheduled TAT'], 'Scheduled TAT (%)')
        
        logger.info(f"Reports exported to {filepath}")
    
    def _format_worksheet(self, worksheet, title: str):
        """Apply formatting to worksheet"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Format header row (now single-level after flattening)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format percentage values (all data cells)
        for row in range(2, worksheet.max_row + 1):
            for col in range(2, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                # Format as percentage with 2 decimal places
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format row headers (client names in first column)
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                      min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, worksheet.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def get_summary(self) -> dict:
        """Get summary statistics"""
        summary = {}
        
        if self.closure_report is not None and not self.closure_report.empty:
            summary['closure_tat_clients'] = len(self.closure_report)
        
        if self.scheduled_report is not None and not self.scheduled_report.empty:
            summary['scheduled_tat_clients'] = len(self.scheduled_report)
        
        return summary


def generate_closure_tat_mis(df: pd.DataFrame) -> tuple:
    """
    Convenience function to generate Closure TAT MIS reports
    
    Args:
        df: Normalized DataFrame
    
    Returns:
        Tuple of (closure_tat_report, scheduled_tat_report)
    """
    generator = ClosureTATMISGenerator(df)
    return generator.generate()
