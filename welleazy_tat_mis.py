"""
TAT MIS Report Logic
Identifies cases with high TAT (Turnaround Time)
"""

import pandas as pd
from datetime import datetime
from welleazy_config import MISConfig
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TATMISGenerator:
    """Generates TAT MIS Report"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        
        # Ensure case_id column exists
        if 'case_id' not in self.df.columns:
            logger.warning("case_id column not found, creating auto-generated IDs")
            self.df['case_id'] = range(1, len(self.df) + 1)
        
        self.report = None
    
    def generate(self) -> pd.DataFrame:
        """
        Generate TAT MIS Report
        
        Process:
        1. Filter for 'Medical Done-Report Awaited' status
        2. Calculate TAT (Current Date - Appointment Date)
        3. Sort by TAT (highest first)
        4. Flag cases with TAT > threshold
        
        Returns:
            DataFrame with TAT analysis
        """
        logger.info("Generating TAT MIS Report...")
        
        # Step 1: Filter by status
        filtered_df = self._filter_by_status()
        
        if filtered_df.empty:
            logger.warning("No cases found with 'Medical Done-Report Awaited' status")
            return pd.DataFrame()
        
        # Step 2: Calculate TAT
        report = self._calculate_tat(filtered_df)
        
        # Step 3: Sort by TAT
        report = self._sort_by_tat(report)
        
        # Step 4: Flag high TAT cases
        report = self._flag_high_tat(report)
        
        # Step 5: Select relevant columns
        report = self._select_columns(report)
        
        self.report = report
        logger.info(f"TAT MIS Report generated. Total cases: {len(report)}, High TAT cases: {report['high_tat'].sum()}")
        
        return report
    
    def _filter_by_status(self) -> pd.DataFrame:
        """Filter for Medical Done-Report Awaited status"""
        target_statuses = MISConfig.get_status_filter('tat_mis')
        
        # Case-insensitive matching
        # Case-insensitive matching - ensure string type
        mask = self.df['case_status'].astype(str).str.lower().isin(
            [s.lower() for s in target_statuses]
        )
        
        filtered = self.df[mask].copy()
        logger.info(f"Filtered {len(filtered)} cases with status: {target_statuses}")
        
        return filtered
    
    def _calculate_tat(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculate TAT in days"""
        current_date = pd.Timestamp.now().normalize()
        
        # Ensure appointment_date is datetime
        df['appointment_date'] = pd.to_datetime(df['appointment_date'], errors='coerce')
        
        # Calculate TAT
        df['tat_days'] = (current_date - df['appointment_date']).dt.days
        
        # Handle negative TATs (future appointments)
        df['tat_days'] = df['tat_days'].clip(lower=0)
        
        logger.info(f"Calculated TAT for {len(df)} cases")
        
        return df
    
    def _sort_by_tat(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sort by TAT descending (highest first)"""
        return df.sort_values('tat_days', ascending=False).reset_index(drop=True)
    
    def _flag_high_tat(self, df: pd.DataFrame) -> pd.DataFrame:
        """Flag cases with TAT above threshold"""
        df['high_tat'] = df['tat_days'] > MISConfig.TAT_THRESHOLD
        return df
    
    def _select_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Select and order relevant columns for report"""
        columns = [
            'client_name',
            'customer_name',
            'case_id',
            'appointment_date',
            'tat_days',
            'high_tat',
            'case_status',
            'dc_name',
            'dc_city'
        ]
        
        # Only select columns that exist
        available_cols = [col for col in columns if col in df.columns]
        
        return df[available_cols]
    
    def get_summary(self) -> dict:
        """Get summary statistics of the TAT report"""
        if self.report is None or self.report.empty:
            return {}
        
        return {
            'total_cases': len(self.report),
            'high_tat_cases': self.report['high_tat'].sum(),
            'avg_tat': self.report['tat_days'].mean(),
            'max_tat': self.report['tat_days'].max(),
            'min_tat': self.report['tat_days'].min(),
            'cases_by_client': self.report.groupby('client_name').size().to_dict()
        }
    
    def export_to_excel(self, filepath: str):
        """Export report to Excel with formatting"""
        if self.report is None or self.report.empty:
            logger.warning("No report to export")
            return
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            self.report.to_excel(writer, sheet_name='TAT MIS', index=False)
            
            # Get worksheet
            worksheet = writer.sheets['TAT MIS']
            
            # Apply formatting
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight high TAT rows
            high_tat_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            for idx, row in self.report.iterrows():
                if row['high_tat']:
                    for cell in worksheet[idx + 2]:  # +2 because of header and 0-indexing
                        cell.fill = high_tat_fill
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, worksheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Report exported to {filepath}")


def generate_tat_mis(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convenience function to generate TAT MIS report
    
    Args:
        df: Normalized DataFrame
    
    Returns:
        TAT MIS report DataFrame
    """
    generator = TATMISGenerator(df)
    return generator.generate()
