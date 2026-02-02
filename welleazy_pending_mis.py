"""
Pending Case MIS Report Logic
Shows count of pending cases by status and date
"""

import pandas as pd
from welleazy_config import MISConfig
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class PendingCaseMISGenerator:
    """Generates Pending Case MIS Report"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        
        # Ensure case_id column exists
        if 'case_id' not in self.df.columns:
            logger.warning("case_id column not found, creating auto-generated IDs")
            self.df['case_id'] = range(1, len(self.df) + 1)
        
        self.report = None
        self.pivot_table = None
    
    def generate(self) -> pd.DataFrame:
        """
        Generate Pending Case MIS Report
        
        Process:
        1. Filter for pending statuses
        2. Create pivot table (Status x Date)
        3. Highlight critical statuses
        
        Returns:
            Pivot table DataFrame
        """
        logger.info("Generating Pending Case MIS Report...")
        
        # Step 1: Filter by pending statuses
        filtered_df = self._filter_pending_cases()
        
        if filtered_df.empty:
            logger.warning("No pending cases found")
            return pd.DataFrame()
        
        # Step 2: Prepare date column
        prepared_df = self._prepare_date_column(filtered_df)
        
        # Step 3: Create pivot table
        pivot = self._create_pivot_table(prepared_df)
        
        # Step 4: Add totals
        pivot = self._add_totals(pivot)
        
        self.pivot_table = pivot
        logger.info(f"Pending Case MIS Report generated. Total cases: {len(filtered_df)}")
        
        return pivot
    
    def _filter_pending_cases(self) -> pd.DataFrame:
        """Filter for pending case statuses"""
        target_statuses = MISConfig.get_status_filter('pending_mis')
        
        # Remove None values first
        self.df = self.df[self.df['case_status'].notna()].copy()
        
        # Case-insensitive matching
        mask = self.df['case_status'].str.lower().isin(
            [s.lower() for s in target_statuses]
        )
        
        filtered = self.df[mask].copy()
        logger.info(f"Filtered {len(filtered)} pending cases")
        
        return filtered
    
    def _prepare_date_column(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare case received date for pivot"""
        # Ensure date is datetime
        df['case_received_date'] = pd.to_datetime(df['case_received_date'], errors='coerce')
        
        # Extract month-year for grouping
        df['month_year'] = df['case_received_date'].dt.to_period('M').astype(str)
        
        # Handle missing dates
        df['month_year'].fillna('Unknown', inplace=True)
        
        return df
    
    def _create_pivot_table(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create pivot table with Status as rows and Month-Year as columns"""
        # Remove any None values from case_status before pivot
        df = df[df['case_status'].notna()].copy()
        
        pivot = pd.pivot_table(
            df,
            values='case_id',
            index='case_status',
            columns='month_year',
            aggfunc='count',
            fill_value=0
        )
        
        # Remove any None/NaN from index
        pivot = pivot.dropna(how='all')
        pivot = pivot[pivot.index.notna()]
        
        # Sort columns chronologically
        try:
            cols = sorted([col for col in pivot.columns if col != 'Unknown'])
            if 'Unknown' in pivot.columns:
                cols.append('Unknown')
            pivot = pivot[cols]
        except:
            pass  # Keep original order if sorting fails
        
        return pivot
    
    def _add_totals(self, pivot: pd.DataFrame) -> pd.DataFrame:
        """Add row and column totals"""
        # Add total column
        pivot['Total'] = pivot.sum(axis=1)
        
        # Add total row
        pivot.loc['Total'] = pivot.sum(axis=0)
        
        return pivot
    
    def get_critical_cases(self) -> pd.DataFrame:
        """Get counts of critical status cases (Fresh case, Waiting for DC confirmation)"""
        if self.pivot_table is None or self.pivot_table.empty:
            return pd.DataFrame()
        
        critical_statuses = MISConfig.HIGHLIGHT_STATUSES
        
        # Filter for critical statuses (case-insensitive)
        mask = self.pivot_table.index.str.lower().isin(
            [s.lower() for s in critical_statuses]
        )
        
        return self.pivot_table[mask]
    
    def export_to_excel(self, filepath: str):
        """Export report to Excel with formatting"""
        if self.pivot_table is None or self.pivot_table.empty:
            logger.warning("No report to export")
            return
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write pivot table
            self.pivot_table.to_excel(writer, sheet_name='Pending Cases', startrow=0)
            
            # Get worksheet
            worksheet = writer.sheets['Pending Cases']
            
            # Apply formatting
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            # Format column headers (first row)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Format row headers (first column)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
            
            # Highlight critical statuses
            highlight_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            critical_statuses = [s.lower() for s in MISConfig.HIGHLIGHT_STATUSES]
            
            for row_idx, row_name in enumerate(self.pivot_table.index, start=2):
                if row_name.lower() in critical_statuses:
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx, column=col).fill = highlight_fill
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, worksheet.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Report exported to {filepath}")
    
    def get_summary(self) -> dict:
        """Get summary statistics"""
        if self.pivot_table is None or self.pivot_table.empty:
            return {}
        
        # Exclude 'Total' row and column for statistics
        data_only = self.pivot_table.drop('Total', axis=0).drop('Total', axis=1)
        
        return {
            'total_pending': int(self.pivot_table.loc['Total', 'Total']),
            'by_status': self.pivot_table['Total'].drop('Total').to_dict(),
            'critical_cases': int(self.get_critical_cases()['Total'].sum()) if not self.get_critical_cases().empty else 0
        }


def generate_pending_mis(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convenience function to generate Pending Case MIS report
    
    Args:
        df: Normalized DataFrame
    
    Returns:
        Pending Case MIS pivot table
    """
    generator = PendingCaseMISGenerator(df)
    return generator.generate()
